import streamlit as st
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook

def write_to_xlsx(data):
    try:
        wb = load_workbook('data_set_2.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Line Num.", "Area of Commodity", "Assm / Part #", "Assembly Component", "Description", "Unit Cost", "QTY", "Material Cost", "Process Cost", "Fastener Cost", "Tooling Cost", "Total Cost", "Details", "Page #", "Subsystem"])

    sheet.append(data)
    wb.save('data_set_2.xlsx')

def main():
    #Line Num.	Area of Commodity	Assm / Part #	Assembly Component	Description	Unit Cost	QTY	Material Cost	Process Cost	Fastener Cost	Tooling Cost	Total Cost	Details	Page #	Subsystem

    st.title("Streamlit Form")
    # Load the data from data_set.xlsx
    data = pd.read_excel('data_set_2.xlsx')

    # Get the current number of records
    line_no = len(data) + 1
    area_of_commodity = st.text_input("Area of Commodity")
    assm_part = st.text_input("Assm / Part #")
    assembly_component = st.text_input("Assembly Component")
    description = st.text_input("Description")
    unit_cost = st.number_input("Unit Cost")
    qty = st.number_input("QTY")
    material_cost = st.number_input("Material Cost")
    process_cost = st.number_input("Process Cost")
    fastener_cost = st.number_input("Fastener Cost")
    tooling_cost = st.number_input("Tooling Cost")
    total_cost = unit_cost * qty + material_cost + process_cost + fastener_cost + tooling_cost
    details = st.text_area("Details")
    pageno = 1
    subsystem = st.selectbox("Subsystem", data['Subsystem'].unique())

    if st.button("Submit"):
        new_row = [line_no, area_of_commodity, assm_part, assembly_component, description, unit_cost, qty, material_cost, process_cost, fastener_cost, tooling_cost, total_cost, details, pageno, subsystem]
        write_to_xlsx(new_row)
        st.success("Data written to data_set_2.xlsx")




if __name__ == "__main__":
    main()
