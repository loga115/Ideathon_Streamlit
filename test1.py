import streamlit as st 
import pandas as pd
import streamlit_authenticator as sat 
import re
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import Workbook, load_workbook

userDetails = pd.DataFrame(columns = ["Username", "Email", "Password"])
cdir = Path.cwd().resolve()
mod_path = (Path(__file__).parent).resolve()

def unameval(uname):
    return bool(re.match(r"^[a-zA-Z0-9_-]{1, 20}$", uname))

def emailval(email):
    return "@" in email and 2 < len(email) < 320 and bool(re.match(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2, 7}\b", email))

def pwordval(pword):
    return 8 < len(pword) < 20 and bool(re.match(r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$", pword))

def register():
    with st.form("register_form"):
        uname = st.text_input("Username")
        pword = st.text_input("Password", type = "password")
        email = st.text_input("Email")
        submit = st.form_submit_button("Register")

        if submit:
            if (unameval(uname) and emailval(email) and pwordval(pword)):
                userDetails = userDetails.append({"Username": uname, "Email": email, "Password": pword}, ignore_index = True)
            
            else:
                st.write("Please re-enter your details.")
    
    return uname, pword, email

def login():
    with st.form("login_form"):
        uname = st.text_input("Username")
        pword = st.text_input("Password", type = "password")
        submit = st.form_submit_button("Login")

        if submit:
            if uname in userDetails["Username"].values:
                user = userDetails[userDetails["Username"] == uname].iloc[0]

                if user["Password"] == pword:
                    st.write("Welcome back!")
                    return True
                
                else:
                    st.write("Incorrect password entered.")

            else:
                st.write("Username not found. Please register before continuing.")
        
        return False
    
def hover(event, ax, pie, fig):
    if event.inaxes == ax:
        for wedge in pie:
            wedge.set_alpha(0.5)
        
        wedge = pie[event.ind[0]]
        wedge.set_alpha(1)
        fig.canvas.draw_idle()
    
def analytics():
    global cdir
    global mod_path

    data = pd.read_excel(mod_path/'data_set.xlsx')
    totalCosts = data.groupby("Subsystem")["Total Cost"].sum()

    fig, ax = plt.subplots()
    pie1 = ax.pie(totalCosts, labels = totalCosts.index, autopct = "%1.1f%%")
    ax.set_title("Subsystem Cost Distribution")

    fig.canvas.mp1_connect('motion_notify_event', hover(ax, pie1, fig))

    if st.button("Show Pie Chart"):
        if st.session_state.get("show_chart"):
            st.session_state["show_chart"] = False
        else:
            st.session_state["show_chart"] = True

        if st.session_state["show_chart"]:
            st.pyplot(fig)
        else:
            st.empty()

    for subsystem in totalCosts.index:
        if st.button(subsystem):
            if st.session_state.get(subsystem):
                st.session_state[subsystem] = False
            else:
                st.session_state[subsystem] = True

            if st.session_state[subsystem]:
                st.write(data[data["Subsystem"] == subsystem])
            else:
                st.empty()

def write_to_xlsx(data):
    try:
        wb = load_workbook("data_set.xlsx")
        sheet = wb.active

    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Line Num.", "Area of Commodity", "Assm / Part #", "Assembly Component", "Description", "Unit Cost", "Qty", "Material Cost", "Process Cost", "Fastener Cost", "Tooling Cost", "Details", "Page #", "Subsystem"])

    sheet.append(data)
    wb.save("data_set.xlsx")

def inputform():
    st.title("Adding elements into the Inventory")
    data = pd.read_excel("data_set.xlsx")

    line_no = len(data) + 1
    area_of_commodity = st.text_input("Area of Commodity")
    assm_part = st.text_input("Assm / Part #")
    assembly_component = st.text_input("Assembly Component")
    description = st.text_input("Description")
    unit_cost = st.number_input("Unit Cost")
    qty = st.number_input("Qty")
    material_cost = st.number_input("Material Cost")
    process_cost = st.number_input("Process Cost")
    tooling_cost = st.number_input("Tooling Cost")
    total_cost = unit_cost*qty + material_cost + process_cost + tooling_cost
    details = st.text_input("Details")
    pageno = 1
    subsystem = st.selectbox("Subsystem", data["Subsystem"].unique())

    if st.button("Submit"):
        new_row = [line_no, area_of_commodity, assm_part, assembly_component, description, unit_cost, qty, material_cost, process_cost, tooling_cost, total_cost, details, pageno, subsystem]
        write_to_xlsx(new_row)
        st.success("Data written to data_set_2.xlsx")

def main():
    st.title("Login/Register")
    global userDetails
    newDataframe = pd.DataFrame({"Username": "a", "Email": "email", "Password": "a"}, index = [0])
    userDetails = pd.concat([userDetails, newDataframe], ignore_index = True)  # Assign the concatenated DataFrame back to userDetails
    l = login()
    if not l:
        st.warning("Please register before you continue")
    
    if st.button("Register"):
        uname, pword, email = register()

        if uname and pword and email and unameval(uname) and emailval(email) and pwordval(pword):
            st.success("You have successfully registered. Please login to continue.")
            # global userDetails
            newDataframe = pd.DataFrame({"Username": uname, "Email": email, "Password": pword}, index = [0])
            userDetails = pd.concat([userDetails, newDataframe], ignore_index = True)  # Assign the concatenated DataFrame back to userDetails
    



main()